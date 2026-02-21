#!/usr/bin/env python3
"""
Matrix Multiplication CUDA Profiling Script

Compiles, runs, and records timing data for all matmul kernel variants.
Sweeps matrix size N from 256 to 4096 (step 128) with M=100 timed iterations.
Outputs results to time.xlsx using openpyxl.

Usage:
    python profile_matmul.py
"""

import subprocess
import os
import sys
import platform

# ============================================================
# Try to import openpyxl for Excel output
# ============================================================
try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill
except ImportError:
    print("Error: openpyxl is required but not installed.")
    print("Install it with:  pip install openpyxl")
    sys.exit(1)

# ============================================================
# Configuration: N range and iteration count
# ============================================================
# Matrix sizes: 256, 384, 512, ..., 4096 (30 values, all multiples of 32)
N_VALUES = list(range(256, 4096 + 1, 128))

# Number of timed kernel launches per (variant, N) combination
M = 100

# ============================================================
# Variant definitions
# Each variant specifies its source file, executable name,
# and how to construct run arguments for a given (N, M).
# ============================================================
VARIANTS = [
    {
        "name": "baseline",
        "dir": "baseline",
        "src": "mmul.cu",
        "exe": "mmul",
        "run_args": lambda N, M: [str(N), str(M)],
    },
    {
        "name": "alignment_naive",
        "dir": "alignment",
        "src": "mmul.cu",
        "exe": "mmul_align",
        "run_args": lambda N, M: [str(N), str(M), "0"],
    },
    {
        "name": "alignment_transposed_a",
        "dir": "alignment",
        "src": "mmul.cu",
        "exe": "mmul_align",
        "run_args": lambda N, M: [str(N), str(M), "1"],
    },
    {
        "name": "alignment_misaligned_b",
        "dir": "alignment",
        "src": "mmul.cu",
        "exe": "mmul_align",
        "run_args": lambda N, M: [str(N), str(M), "2"],
    },
    {
        "name": "tiled_1d",
        "dir": "tiled",
        "src": "mmul.cu",
        "exe": "mmul_tiled",
        "run_args": lambda N, M: [str(N), str(M)],
    },
    {
        "name": "tiled_2d",
        "dir": "tiled",
        "src": "mmul_2D.cu",
        "exe": "mmul_tiled_2D",
        "run_args": lambda N, M: [str(N), str(M)],
    },
]


def get_exe_suffix():
    """Return '.exe' on Windows, '' on Linux/WSL."""
    if platform.system() == "Windows":
        return ".exe"
    return ""


def compile_variant(base_dir, variant):
    """
    Compile a single variant using nvcc.
    Returns True on success, False on failure.
    """
    src_path = os.path.join(base_dir, variant["dir"], variant["src"])
    exe_path = os.path.join(base_dir, variant["dir"], variant["exe"] + get_exe_suffix())
    cmd = ["nvcc", src_path, "-o", exe_path]

    print(f"  Compiling {variant['name']}... ", end="", flush=True)
    result = subprocess.run(cmd, capture_output=True, text=True)
    if result.returncode != 0:
        print(f"FAILED")
        print(f"  stderr: {result.stderr.strip()}")
        return False
    print("OK")
    return True


def run_variant(base_dir, variant, N, M):
    """
    Run a compiled variant with given N and M.
    Returns the average kernel time in ms (float), or None on error.
    """
    exe_path = os.path.join(base_dir, variant["dir"], variant["exe"] + get_exe_suffix())
    args = variant["run_args"](N, M)
    cmd = [exe_path] + args

    try:
        # Timeout: 10 minutes max per run (large N with M=100 can be slow)
        result = subprocess.run(cmd, capture_output=True, text=True, timeout=600)
        if result.returncode != 0:
            return None
        # Parse the single float output (average time in ms)
        return float(result.stdout.strip().split('\n')[-1])
    except (subprocess.TimeoutExpired, ValueError, IndexError):
        return None


def main():
    # Determine base directory (where this script lives)
    base_dir = os.path.dirname(os.path.abspath(__file__))

    # ========================================================
    # Step 1: Compile all variants
    # ========================================================
    print("=" * 60)
    print("Step 1: Compiling all variants")
    print("=" * 60)

    # Track which source files have been compiled to avoid duplicates
    compiled_sources = set()
    for v in VARIANTS:
        key = os.path.join(v["dir"], v["src"], v["exe"])
        if key in compiled_sources:
            print(f"  {v['name']} (reusing compiled {v['exe']})")
            continue
        if not compile_variant(base_dir, v):
            print(f"\nCompilation failed for {v['name']}. Aborting.")
            sys.exit(1)
        compiled_sources.add(key)

    # ========================================================
    # Step 2: Run profiling sweep
    # ========================================================
    print()
    print("=" * 60)
    print(f"Step 2: Profiling (M={M} iterations per configuration)")
    print("=" * 60)

    # results[variant_name][N] = time_ms
    results = {v["name"]: {} for v in VARIANTS}

    for N in N_VALUES:
        print(f"\nN = {N}:")
        for v in VARIANTS:
            print(f"  {v['name']:30s} ... ", end="", flush=True)
            time_ms = run_variant(base_dir, v, N, M)
            if time_ms is not None:
                results[v["name"]][N] = time_ms
                print(f"{time_ms:.4f} ms")
            else:
                results[v["name"]][N] = None
                print("ERROR")

    # ========================================================
    # Step 3: Write results to Excel
    # ========================================================
    print()
    print("=" * 60)
    print("Step 3: Writing results to time.xlsx")
    print("=" * 60)

    wb = Workbook()
    ws = wb.active
    ws.title = "MatMul Kernel Timing"

    # Header row: N | variant1 | variant2 | ...
    variant_names = [v["name"] for v in VARIANTS]
    header = ["N (matrix size)"] + variant_names
    ws.append(header)

    # Style the header row
    header_font = Font(bold=True)
    header_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
    for col_idx, cell in enumerate(ws[1], 1):
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    # Data rows
    for N in N_VALUES:
        row = [N]
        for vn in variant_names:
            val = results[vn].get(N)
            if val is not None:
                row.append(round(val, 4))
            else:
                row.append("N/A")
        ws.append(row)

    # Auto-adjust column widths
    for col in ws.columns:
        max_len = max(len(str(cell.value)) for cell in col)
        col_letter = col[0].column_letter
        ws.column_dimensions[col_letter].width = max(max_len + 2, 12)

    # Add a metadata row
    ws.append([])
    ws.append([f"M = {M} iterations per measurement. Times in milliseconds (ms)."])

    # Save the workbook
    output_path = os.path.join(base_dir, "time.xlsx")
    wb.save(output_path)
    print(f"Results saved to: {output_path}")
    print("Done!")


if __name__ == "__main__":
    main()
