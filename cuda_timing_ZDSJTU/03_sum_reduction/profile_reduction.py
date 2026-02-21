#!/usr/bin/env python3
"""
Sum Reduction CUDA Profiling Script

Compiles, runs, and records timing data for all reduction kernel variants.
Sweeps array size N = 256 * 2^i (i=0..12) with M=100 timed iterations.
Outputs results to time.xlsx using openpyxl.

Usage:
    python profile_reduction.py
"""

import subprocess
import os
import sys
import platform

# ============================================================
# Try to import openpyxl for Excel output
# ============================================================
try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill
except ImportError:
    print("Error: openpyxl is required but not installed.")
    print("Install it with:  pip install openpyxl")
    sys.exit(1)

# ============================================================
# Configuration: N range and iteration count
# ============================================================
# Array sizes: N = 256 * 2^i, i=0..12
# Values: 256, 512, 1024, 2048, 4096, 8192, 16384, 32768, 65536,
#         131072, 262144, 524288, 1048576
N_VALUES = [256 * (2 ** i) for i in range(13)]

# Number of timed kernel launches per (variant, N) combination
M = 100

# ============================================================
# Variant definitions
# Each variant specifies its source file, executable name,
# minimum valid N, and how to construct run arguments.
# ============================================================
VARIANTS = [
    {
        "name": "diverged",
        "dir": "diverged",
        "src": "sumReduction.cu",
        "exe": "sumReduction",
        "min_N": 256,   # Single-load: need N >= TB_SIZE
        "run_args": lambda N, M: [str(N), str(M)],
    },
    {
        "name": "bank_conflicts",
        "dir": "bank_conflicts",
        "src": "sumReduction.cu",
        "exe": "sumReduction",
        "min_N": 256,
        "run_args": lambda N, M: [str(N), str(M)],
    },
    {
        "name": "no_conflicts",
        "dir": "no_conflicts",
        "src": "sumReduction.cu",
        "exe": "sumReduction",
        "min_N": 256,
        "run_args": lambda N, M: [str(N), str(M)],
    },
    {
        "name": "reduce_idle",
        "dir": "reduce_idle",
        "src": "sumReduction.cu",
        "exe": "sumReduction",
        "min_N": 512,   # Double-load: need N >= TB_SIZE * 2
        "run_args": lambda N, M: [str(N), str(M)],
    },
    {
        "name": "device_function",
        "dir": "device_function",
        "src": "sumReduction.cu",
        "exe": "sumReduction",
        "min_N": 512,   # Double-load: need N >= TB_SIZE * 2
        "run_args": lambda N, M: [str(N), str(M)],
    },
    {
        "name": "cooperative_groups",
        "dir": "cooperative_groups",
        "src": "sumReduction.cu",
        "exe": "sumReduction",
        "min_N": 256,   # Grid-stride loop handles any size
        "run_args": lambda N, M: [str(N), str(M)],
    },
]


def get_exe_suffix():
    """Return '.exe' on Windows, '' on Linux/WSL."""
    if platform.system() == "Windows":
        return ".exe"
    return ""


def compile_variant(base_dir, variant):
    """
    Compile a single variant using nvcc.
    Returns True on success, False on failure.
    """
    src_path = os.path.join(base_dir, variant["dir"], variant["src"])
    exe_path = os.path.join(base_dir, variant["dir"], variant["exe"] + get_exe_suffix())
    cmd = ["nvcc", src_path, "-o", exe_path]

    print(f"  Compiling {variant['name']}... ", end="", flush=True)
    result = subprocess.run(cmd, capture_output=True, text=True)
    if result.returncode != 0:
        print(f"FAILED")
        print(f"  stderr: {result.stderr.strip()}")
        return False
    print("OK")
    return True


def run_variant(base_dir, variant, N, M):
    """
    Run a compiled variant with given N and M.
    Returns the average kernel time in ms (float), or None on error.
    """
    exe_path = os.path.join(base_dir, variant["dir"], variant["exe"] + get_exe_suffix())
    args = variant["run_args"](N, M)
    cmd = [exe_path] + args

    try:
        # Timeout: 5 minutes max per run
        result = subprocess.run(cmd, capture_output=True, text=True, timeout=300)
        if result.returncode != 0:
            return None
        # Parse the single float output (average time in ms)
        return float(result.stdout.strip().split('\n')[-1])
    except (subprocess.TimeoutExpired, ValueError, IndexError):
        return None


def main():
    # Determine base directory (where this script lives)
    base_dir = os.path.dirname(os.path.abspath(__file__))

    # ========================================================
    # Step 1: Compile all variants
    # ========================================================
    print("=" * 60)
    print("Step 1: Compiling all variants")
    print("=" * 60)

    for v in VARIANTS:
        if not compile_variant(base_dir, v):
            print(f"\nCompilation failed for {v['name']}. Aborting.")
            sys.exit(1)

    # ========================================================
    # Step 2: Run profiling sweep
    # ========================================================
    print()
    print("=" * 60)
    print(f"Step 2: Profiling (M={M} iterations per configuration)")
    print("=" * 60)

    # results[variant_name][N] = time_ms
    results = {v["name"]: {} for v in VARIANTS}

    for N in N_VALUES:
        print(f"\nN = {N:>10d}:")
        for v in VARIANTS:
            # Skip variants that cannot handle this N
            if N < v["min_N"]:
                print(f"  {v['name']:25s} ... SKIPPED (N < {v['min_N']})")
                results[v["name"]][N] = None
                continue

            print(f"  {v['name']:25s} ... ", end="", flush=True)
            time_ms = run_variant(base_dir, v, N, M)
            if time_ms is not None:
                results[v["name"]][N] = time_ms
                print(f"{time_ms:.6f} ms")
            else:
                results[v["name"]][N] = None
                print("ERROR")

    # ========================================================
    # Step 3: Write results to Excel
    # ========================================================
    print()
    print("=" * 60)
    print("Step 3: Writing results to time.xlsx")
    print("=" * 60)

    wb = Workbook()
    ws = wb.active
    ws.title = "Sum Reduction Timing"

    # Header row: N | variant1 | variant2 | ...
    variant_names = [v["name"] for v in VARIANTS]
    header = ["N (array size)"] + variant_names
    ws.append(header)

    # Style the header row
    header_font = Font(bold=True)
    header_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
    for col_idx, cell in enumerate(ws[1], 1):
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    # Data rows
    for N in N_VALUES:
        row = [N]
        for vn in variant_names:
            val = results[vn].get(N)
            if val is not None:
                row.append(round(val, 6))
            else:
                row.append("N/A")
        ws.append(row)

    # Auto-adjust column widths
    for col in ws.columns:
        max_len = max(len(str(cell.value)) for cell in col)
        col_letter = col[0].column_letter
        ws.column_dimensions[col_letter].width = max(max_len + 2, 14)

    # Add metadata
    ws.append([])
    ws.append([f"M = {M} iterations per measurement. Times in milliseconds (ms)."])
    ws.append(["reduce_idle and device_function require N >= 512 (double-load)."])

    # Save the workbook
    output_path = os.path.join(base_dir, "time.xlsx")
    wb.save(output_path)
    print(f"Results saved to: {output_path}")
    print("Done!")


if __name__ == "__main__":
    main()
